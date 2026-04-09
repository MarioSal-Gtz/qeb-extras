"""
Fill missing CSV fields (Articulo, Precio por cara, Operacion) from SAP Orders.
APS Global = DocNum.
- APS 1000-2000 -> SBOIMUTRADE (v2)
- APS 60000+    -> SBOCIMU (v1)
Matches DocumentLine by city code derived from Unidad prefix.
Cross-references with inventory Excel to validate.
"""

import csv
import json
import re
import time
import urllib.request
import urllib.error
import ssl

# === CONFIG ===
SAP_DB_TRADE = {"name": "SBOIMUTRADE", "base": "https://10.1.0.9:50000/b1s/v2", "CompanyDB": "SBOIMUTRADE", "UserName": "manager", "Password": "Grup0$1Mu$"}
SAP_DB_CIMU = {"name": "SBOCIMU", "base": "https://10.1.0.9:50000/b1s/v1", "CompanyDB": "SBOCIMU", "UserName": "manager", "Password": "Grup0$1Mu$"}

CSV_INPUT = r"C:\Users\Mario\Downloads\Layout QEB - Cat08-Cat26 - SIN DIDI.csv"
CSV_OUTPUT = r"C:\Users\Mario\Downloads\Layout QEB - Cat08-Cat26 - FILLED.csv"
INVENTORY_XLSX = r"C:\Users\Mario\Downloads\Inventario Grupo IMU - QEB 02.03.2026 (1).xlsx"

ctx = ssl.create_default_context()
ctx.check_hostname = False
ctx.verify_mode = ssl.CERT_NONE

# === UNIDAD PREFIX -> CITY CODE ===
UNIDAD_PREFIX_TO_CITY = {
    "AC": "AC",
    "AL": "BR",
    "ATZP": "MX",
    "BR": "BR",
    "CH": "PU",
    "CIZC": "MX",
    "CU": "CU",
    "DF": "MX",
    "ECAT": "MX",
    "ED": "MX",
    "HUIX": "MX",
    "LEO": "LE",
    "LERM": "TL",
    "MD": "MR",
    "METE": "TL",
    "MS": "MX",
    "MZ": "MZ",
    "NAUC": "MX",
    "OX": "OX",
    "PH": "PH",
    "PUE": "PU",
    "PV": "PV",
    "SMAT": "TL",
    "TJ": "TJ",
    "TLAN": "MX",
    "TOLU": "TL",
    "TULT": "MX",
    "VR": "BR",
    "ZP": "GD",
    "Z": "GD",
}

# Ciudad name -> city code fallback
CIUDAD_TO_CITY = {
    "Acapulco de Juárez": "AC",
    "Acapulco": "AC",
    "Alvarado": "BR",
    "Boca del Río": "BR",
    "Ciudad de México": "MX",
    "Culiacán": "CU",
    "Guadalajara": "GD",
    "Guadalupe": "MY",
    "León": "LE",
    "Mazatlán": "MZ",
    "Monterrey": "MY",
    "Mérida": "MR",
    "Oaxaca de Juárez": "OX",
    "Oaxaca": "OX",
    "Pachuca de Soto": "PH",
    "Pachuca": "PH",
    "Puebla": "PU",
    "Puerto Vallarta": "PV",
    "San Nicolás de los Garza": "MY",
    "San Pedro Tlaquepaque": "GD",
    "Tijuana": "TJ",
    "Toluca": "TL",
    "Veracruz": "BR",
    "Zapopan": "GD",
    "Estado de México": "MX",
    "Aguascalientes": "AG",
    "San Luis Potosí": "SL",
}

# City code aliases: SAP article city code -> our normalized city code
# e.g. SAP uses PB for Puebla but we use PU
CITY_CODE_ALIASES = {
    "PB": "PU",
    "VE": "BR",  # Veracruz
}

# Operacion mapping from ItemCode prefix
PREFIX_TO_OP = {
    "RT": "RENTA",
    "BF": "BONIFICACION",
    "CT": "CORTESIA",
    "IN": "INTERCAMBIO",
}

# CF-* legacy articles -> (city_code, operacion)
CF_MAP = {
    "CF-0001": ("MX", "RENTA"), "CF-0002": ("MX", "RENTA"), "CF-0003": ("MX", "RENTA"),
    "CF-0004": ("MX", "RENTA"), "CF-0005": ("MX", "RENTA"), "CF-0007": ("MX", "RENTA"),
    "CF-0008": ("MX", "RENTA"), "CF-0009": ("MX", "RENTA"), "CF-0011": ("MX", "RENTA"),
    "CF-0012": ("MX", "RENTA"), "CF-0013": ("PH", "RENTA"), "CF-0014": ("TJ", "RENTA"),
    "CF-0015": ("AC", "RENTA"), "CF-0016": ("OX", "RENTA"), "CF-0017": ("PV", "RENTA"),
    "CF-0018": ("MZ", "RENTA"), "CF-0020": ("BR", "RENTA"), "CF-0021": ("GD", "RENTA"),
    "CF-0022": ("MR", "RENTA"), "CF-0010": ("MX", "RENTA"),
    "CF-0023": ("MX", "RENTA"),
    "CF-0024": ("MX", "RENTA"),
    "CF-0025": ("MX", "RENTA"),
    "CF-0026": ("MX", "RENTA"),
    "CF-0030": ("MX", "BONIFICACION"), "CF-0031": ("MX", "BONIFICACION"),
    "CF-0038": ("MX", "BONIFICACION"),
    "CF-0043": ("MX", "BONIFICACION"), "CF-0044": ("MX", "BONIFICACION"),
    "CF-0045": ("PH", "BONIFICACION"), "CF-0046": ("TJ", "BONIFICACION"),
    "CF-0047": ("AC", "BONIFICACION"), "CF-0048": ("OX", "BONIFICACION"),
    "CF-0049": ("PV", "BONIFICACION"), "CF-0051": ("BR", "BONIFICACION"),
    "CF-0052": ("GD", "BONIFICACION"), "CF-0053": ("MR", "BONIFICACION"),
    "CF-0054": ("MX", "BONIFICACION"), "CF-0055": ("MX", "BONIFICACION"),
    "CF-0056": ("MX", "BONIFICACION"), "CF-0094": ("MX", "BONIFICACION"),
    "CF-0096": ("MZ", "BONIFICACION"),
    "CF-0039": ("MX", "CORTESIA"), "CF-0040": ("MX", "CORTESIA"),
    "CF-0057": ("MX", "CORTESIA"), "CF-0058": ("MX", "CORTESIA"),
    "CF-0059": ("PH", "CORTESIA"), "CF-0060": ("TJ", "CORTESIA"),
    "CF-0061": ("AC", "CORTESIA"), "CF-0062": ("OX", "CORTESIA"),
    "CF-0063": ("PV", "CORTESIA"), "CF-0064": ("MZ", "CORTESIA"),
    "CF-0065": ("BR", "CORTESIA"), "CF-0066": ("GD", "CORTESIA"),
    "CF-0067": ("MR", "CORTESIA"), "CF-0068": ("MX", "CORTESIA"),
    "CF-0069": ("MX", "CORTESIA"), "CF-0070": ("MX", "CORTESIA"),
    "CF-0095": ("MX", "RENTA"),
    "CF-0097": ("AC", "RENTA"), "CF-0098": ("MZ", "RENTA"),
    "CF-0099": ("PH", "RENTA"), "CF-0104": ("TJ", "RENTA"),
    "CF-0107": ("MX", "RENTA"),
    "CF-0108": ("AC", "BONIFICACION"), "CF-0109": ("MZ", "BONIFICACION"),
    "CF-0110": ("PH", "BONIFICACION"), "CF-0111": ("TJ", "BONIFICACION"),
    "CF-0118": ("MX", "BONIFICACION"), "CF-0119": ("AC", "CORTESIA"),
}


# Inventory mueble -> SAP article type mapping
# SAP articles: RT-P1-COB-MX (P1=parabus), RT-CL-COB-MX (CL=columna), etc.
MUEBLE_TO_ART_TYPE = {
    "PARABUS": ["P1", "P2", "P4"],
    "COLUMNA": ["CL"],
    "KIOSCO": ["KCS"],
    "BOLERO PARABUS": ["BL"],
    "BOLERO": ["BL"],
    "MUPI": ["P1", "P2"],  # MUPIs use parabus articles
    "PUENTE PEATONAL": ["P1"],
    "BAJO PUENTES": ["P1"],
    "MULTISERVICIO": ["P1"],
    "UNIPOLAR": ["P1"],
    "PARATAXI": ["P1"],
}


def load_inventory():
    """Load inventory Excel and build a dict: cod -> {plaza, mueble, tipo_de_cara, ...}."""
    try:
        import openpyxl
    except ImportError:
        print("  openpyxl not installed, skipping inventory cross-reference")
        return {}

    print(f"Loading inventory: {INVENTORY_XLSX}")
    wb = openpyxl.load_workbook(INVENTORY_XLSX, read_only=True)
    ws = wb[wb.sheetnames[0]]

    inventory = {}
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i <= 1:
            continue
        # Columns: Cod(0), codigo_unico(1), ubicacion(2), tipo_de_cara(3), cara(4), mueble(5),
        #          lat(6), lon(7), plaza(8), estado(9), municipio(10), ...
        #          tipo_de_mueble(17)
        cod = str(row[0]).strip() if row[0] else ""
        plaza = str(row[8]).strip() if row[8] else ""
        mueble = str(row[5]).strip().upper() if row[5] else ""
        tipo_de_cara = str(row[3]).strip() if row[3] else ""
        tradicional_digital = str(row[12]).strip() if row[12] else ""

        if cod:
            inventory[cod] = {
                "plaza": plaza,
                "mueble": mueble,
                "tipo_de_cara": tipo_de_cara,
                "tradicional_digital": tradicional_digital,
            }

    wb.close()
    print(f"  Loaded {len(inventory)} inventory items")
    return inventory


def get_city_code(unidad, ciudad):
    """Get city code from Unidad prefix, fallback to Ciudad name."""
    if unidad:
        prefix = re.match(r'^([A-Za-z]+)', unidad)
        if prefix:
            p = prefix.group(1).upper()
            for l in range(len(p), 0, -1):
                if p[:l] in UNIDAD_PREFIX_TO_CITY:
                    return UNIDAD_PREFIX_TO_CITY[p[:l]]

    # Fallback to ciudad name
    if ciudad:
        for name, code in CIUDAD_TO_CITY.items():
            if name.lower() in ciudad.lower() or ciudad.lower() in name.lower():
                return code

    return None


def normalize_city_code(code):
    """Normalize SAP article city codes using aliases."""
    return CITY_CODE_ALIASES.get(code, code)


def sap_request(url, method="GET", data=None, cookie=None):
    headers = {"Content-Type": "application/json"}
    if cookie:
        headers["Cookie"] = cookie
    body = json.dumps(data).encode() if data else None
    req = urllib.request.Request(url, data=body, headers=headers, method=method)
    resp = urllib.request.urlopen(req, context=ctx, timeout=30)
    return json.loads(resp.read().decode())


def sap_login(db):
    url = f"{db['base']}/Login"
    login_data = {"CompanyDB": db["CompanyDB"], "UserName": db["UserName"], "Password": db["Password"]}
    result = sap_request(url, method="POST", data=login_data)
    session_id = result.get("SessionId", "")
    print(f"  Logged in to {db['name']}")
    return f"B1SESSION={session_id}"


def get_order_by_docnum(docnum, base_url, cookie):
    url = (
        f"{base_url}/Orders?$filter=DocNum%20ge%20{docnum}%20and%20DocNum%20le%20{docnum}"
        f"&$select=DocEntry,DocNum,CardCode,CardName,DocumentLines"
    )
    result = sap_request(url, cookie=cookie)
    values = result.get("value", [])
    return values[0] if values else None


def parse_order_lines(order):
    """Parse DocumentLines into a dict: city_code -> list of {item_code, prefix, price, op}."""
    if not order:
        return {}
    lines = order.get("DocumentLines", [])
    city_lines = {}

    for line in lines:
        ic = line.get("ItemCode", "")
        price = line.get("Price", 0)

        # Check CF-* legacy articles first
        if ic in CF_MAP:
            city_code, op = CF_MAP[ic]
            if city_code not in city_lines:
                city_lines[city_code] = []
            city_lines[city_code].append({
                "item_code": ic,
                "prefix": "CF",
                "price": price,
                "op": op,
            })
            continue

        # Parse ItemCode: PREFIX-TYPE-FORMAT-CITY (e.g., RT-P1-COB-MX)
        parts = ic.split("-")
        if len(parts) < 4:
            continue

        prefix = parts[0]  # RT, BF, CT, IN, PQ, IM, ES
        art_type = parts[1]  # P1, P2, CL, KCS, BL, DIG, etc.
        raw_city = parts[-1]  # MX, AC, GD, PB, etc.
        city_code = normalize_city_code(raw_city)  # PB -> PU, etc.

        # Skip non-relevant prefixes
        if prefix not in PREFIX_TO_OP:
            continue

        op = PREFIX_TO_OP[prefix]

        if city_code not in city_lines:
            city_lines[city_code] = []
        city_lines[city_code].append({
            "item_code": ic,
            "prefix": prefix,
            "art_type": art_type,  # P1, CL, KCS, BL, DIG, etc.
            "price": price,
            "op": op,
        })

    return city_lines


def pick_line_for_city(city_lines, city_code, mueble=None):
    """Pick the best DocumentLine for a city, matching mueble type from inventory.
    Priority: RT > BF > CT > IN. If mueble is known, filter to matching art_types first."""
    lines = city_lines.get(city_code, [])
    if not lines:
        return None

    # If we know the mueble, try to filter lines to matching article types
    if mueble:
        allowed_types = MUEBLE_TO_ART_TYPE.get(mueble)
        if allowed_types:
            matching = [l for l in lines if l.get("art_type") in allowed_types or l.get("prefix") == "CF"]
            if matching:
                lines = matching
            # If no matching lines, fall through to use all lines (will be flagged as REVISAR)

    def get_priority(x):
        if x["prefix"] == "CF":
            op_priority = {"RENTA": 0, "BONIFICACION": 1, "CORTESIA": 2, "INTERCAMBIO": 3}
            return op_priority.get(x["op"], 99)
        return {"RT": 0, "BF": 1, "CT": 2, "IN": 3}.get(x["prefix"], 99)
    lines.sort(key=get_priority)
    return lines[0]


def format_price(price, op):
    """Format price like Cat05: $X,XXX or $0."""
    if op in ("BONIFICACION", "CORTESIA", "INTERCAMBIO") or not price:
        return "$0"
    p = int(round(price))
    return f"${p:,}"


def get_db_for_aps(aps_num):
    """Determine which DB to query based on APS range."""
    if 1000 <= aps_num <= 2000:
        return "SBOIMUTRADE"
    elif aps_num >= 60000:
        return "SBOCIMU"
    else:
        # Try both
        return "BOTH"


def main():
    print("=== Fill CSV from SAP Orders (v3) ===\n")

    # Load inventory for cross-reference
    inventory = load_inventory()

    # Login to both databases
    sessions = {}
    for db in [SAP_DB_TRADE, SAP_DB_CIMU]:
        try:
            sessions[db["name"]] = {"cookie": sap_login(db), "base": db["base"], "db": db}
        except Exception as e:
            print(f"  FAILED: {db['name']}: {e}")

    # Read CSV
    print(f"\nReading CSV: {CSV_INPUT}")
    with open(CSV_INPUT, "r", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        fieldnames = list(reader.fieldnames) if reader.fieldnames else []
        all_rows = list(reader)
    print(f"Total rows: {len(all_rows)}")

    # Add Notas column for flags if not present
    if "Notas" not in fieldnames:
        fieldnames.append("Notas")

    # Filter out empty rows
    rows = []
    empty_rows = 0
    for r in all_rows:
        campana = (r.get("Campaña") or "").strip()
        unidad = (r.get("Unidad") or "").strip()
        if campana or unidad:
            rows.append(r)
        else:
            empty_rows += 1
    print(f"Non-empty rows: {len(rows)}, Empty rows removed: {empty_rows}")

    # Collect unique APS values and classify by DB
    aps_set = set()
    for r in rows:
        aps = (r.get("APS Global") or "").strip()
        if aps:
            aps_set.add(aps)

    # Separate APS by DB range
    aps_trade = []
    aps_cimu = []
    aps_both = []
    for aps in sorted(aps_set, key=lambda x: int(x) if x.isdigit() else 0):
        aps_num = int(aps) if aps.isdigit() else 0
        db_target = get_db_for_aps(aps_num)
        if db_target == "SBOIMUTRADE":
            aps_trade.append(aps)
        elif db_target == "SBOCIMU":
            aps_cimu.append(aps)
        else:
            aps_both.append(aps)

    print(f"\nUnique APS: {len(aps_set)}")
    print(f"  TRADE (1000-2000): {len(aps_trade)}")
    print(f"  CIMU (60000+): {len(aps_cimu)}")
    print(f"  Unknown range (try both): {len(aps_both)}")

    # Query SAP
    aps_data = {}

    def query_aps(aps, db_name, sess, label):
        try:
            order = get_order_by_docnum(aps, sess["base"], sess["cookie"])
            if order:
                city_lines = parse_order_lines(order)
                aps_data[aps] = {
                    "order": order,
                    "city_lines": city_lines,
                    "db": db_name,
                    "card_name": order.get("CardName", ""),
                }
                cities_str = ",".join(sorted(city_lines.keys()))
                print(f"  {label} APS {aps}: {db_name} - {order.get('CardName','')} | Cities: {cities_str}")
                return True
        except Exception as e:
            if "401" in str(e) or "session" in str(e).lower():
                try:
                    sess["cookie"] = sap_login(sess["db"])
                    order = get_order_by_docnum(aps, sess["base"], sess["cookie"])
                    if order:
                        city_lines = parse_order_lines(order)
                        aps_data[aps] = {
                            "order": order,
                            "city_lines": city_lines,
                            "db": db_name,
                            "card_name": order.get("CardName", ""),
                        }
                        return True
                except:
                    pass
        return False

    # Query CIMU first (60000+)
    if aps_cimu and "SBOCIMU" in sessions:
        print(f"\n--- Querying SBOCIMU ({len(aps_cimu)} APS) ---")
        for i, aps in enumerate(aps_cimu):
            label = f"[{i+1}/{len(aps_cimu)}]"
            if not query_aps(aps, "SBOCIMU", sessions["SBOCIMU"], label):
                print(f"  {label} APS {aps}: NOT FOUND in SBOCIMU")
            time.sleep(0.15)

    # Query TRADE (1000-2000)
    if aps_trade and "SBOIMUTRADE" in sessions:
        print(f"\n--- Querying SBOIMUTRADE ({len(aps_trade)} APS) ---")
        for i, aps in enumerate(aps_trade):
            label = f"[{i+1}/{len(aps_trade)}]"
            if not query_aps(aps, "SBOIMUTRADE", sessions["SBOIMUTRADE"], label):
                # Try CIMU as fallback
                if "SBOCIMU" in sessions:
                    if not query_aps(aps, "SBOCIMU", sessions["SBOCIMU"], label):
                        print(f"  {label} APS {aps}: NOT FOUND")
                else:
                    print(f"  {label} APS {aps}: NOT FOUND in SBOIMUTRADE")
            time.sleep(0.15)

    # Query unknown range (try both)
    if aps_both:
        print(f"\n--- Querying unknown range ({len(aps_both)} APS, trying both DBs) ---")
        for i, aps in enumerate(aps_both):
            label = f"[{i+1}/{len(aps_both)}]"
            found = False
            for db_name in ["SBOIMUTRADE", "SBOCIMU"]:
                if db_name in sessions:
                    if query_aps(aps, db_name, sessions[db_name], label):
                        found = True
                        break
                    time.sleep(0.1)
            if not found:
                print(f"  {label} APS {aps}: NOT FOUND in any DB")
            time.sleep(0.15)

    # Fill CSV rows
    print(f"\nFilling CSV rows...")
    filled = 0
    no_match = 0
    no_aps = 0
    revisar = 0

    for r in rows:
        aps = (r.get("APS Global") or "").strip()
        if not aps or aps not in aps_data:
            if not aps:
                no_aps += 1
            continue

        data = aps_data[aps]
        city_lines = data["city_lines"]

        unidad = (r.get("Unidad") or "").strip()
        ciudad = (r.get("Ciudad") or "").strip()
        city_code = get_city_code(unidad, ciudad)

        if not city_code:
            no_match += 1
            r["Notas"] = "REVISAR: no se pudo determinar ciudad"
            revisar += 1
            continue

        # Get mueble type from inventory to pick correct article
        mueble = None
        inv_item = None
        if inventory and unidad:
            inv_item = inventory.get(unidad)
            if inv_item:
                mueble = inv_item["mueble"]

        line = pick_line_for_city(city_lines, city_code, mueble=mueble)
        if not line:
            no_match += 1
            available_cities = ",".join(sorted(city_lines.keys()))
            r["Notas"] = f"REVISAR: APS no tiene articulo para {city_code} (tiene: {available_cities})"
            revisar += 1
            continue

        # Fill fields
        if not (r.get("Articulo") or "").strip():
            r["Articulo"] = line["item_code"]

        op_key = "Operación" if "Operación" in r else "Operacion"
        if not (r.get(op_key) or "").strip():
            r[op_key] = line["op"]

        precio_key = "Precio por cara (Opcional)" if "Precio por cara (Opcional)" in r else "Precio por cara"
        current_precio = (r.get(precio_key) or "").strip()
        if not current_precio or current_precio == "0":
            r[precio_key] = format_price(line["price"], line["op"])

        # Validate: check if assigned article type matches inventory mueble
        notas = []
        if inv_item:
            # Check mueble vs article type match
            if mueble and line.get("art_type"):
                allowed = MUEBLE_TO_ART_TYPE.get(mueble)
                if allowed and line["art_type"] not in allowed:
                    notas.append(f"REVISAR: mueble={mueble} pero articulo tipo={line['art_type']} ({line['item_code']})")

            # Check inventory plaza vs city code
            inv_plaza = inv_item["plaza"]
            inv_city = None
            for name, code in CIUDAD_TO_CITY.items():
                if name.lower() in inv_plaza.lower() or inv_plaza.lower() in name.lower():
                    inv_city = code
                    break
            if inv_city and inv_city != city_code:
                notas.append(f"REVISAR: inventario plaza={inv_plaza} ({inv_city}) != ciudad {city_code}")
        elif inventory and unidad:
            notas.append(f"REVISAR: Unidad {unidad} no encontrada en inventario")

        if notas:
            r["Notas"] = " | ".join(notas)
            revisar += 1

        filled += 1

    # Write output
    print(f"\nWriting output CSV: {CSV_OUTPUT}")
    with open(CSV_OUTPUT, "w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)

    print(f"\nDone!")
    print(f"  Rows filled: {filled}")
    print(f"  Rows without APS: {no_aps}")
    print(f"  Rows with APS but no city match: {no_match}")
    print(f"  Rows marked REVISAR: {revisar}")
    print(f"  Empty rows removed: {empty_rows}")
    print(f"  Final row count: {len(rows)}")

    # Summary per APS
    print(f"\n=== APS Summary ===")
    all_sorted = sorted(aps_data.keys(), key=lambda x: int(x) if x.isdigit() else 0)
    for aps in all_sorted:
        d = aps_data[aps]
        cl = d["city_lines"]
        cities_detail = []
        for city, lines in sorted(cl.items()):
            items = [f"{l['item_code']}(${l['price']})" for l in lines]
            cities_detail.append(f"{city}:[{','.join(items)}]")
        print(f"  APS {aps}: [{d['db']}] {d['card_name']}")
        print(f"    {' | '.join(cities_detail)}")

    not_found = [a for a in sorted(aps_set, key=lambda x: int(x) if x.isdigit() else 0) if a not in aps_data]
    if not_found:
        print(f"\n  NOT FOUND in any DB: {', '.join(not_found)}")


if __name__ == "__main__":
    main()
